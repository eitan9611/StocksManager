from dialogs import *
from cards import *
import sys
import os
from main import *
import pandas as pd
from PyQt5.QtWidgets import QFileDialog
import os
from pathlib import Path
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import numpy as np
from datetime import *

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from Present.TradePresent import *
from Present.UserPresent import *
from Present.StockPresent import *

# Import proper SVG handling
from PyQt5.QtSvg import QSvgRenderer
from PyQt5.QtGui import QPixmap, QPainter



class BodyContentFrame(ScrollableFrame):
    email = ""
    def __init__(self,email):
        super().__init__()


        self.setStyleSheet("""  
        QScrollBar:vertical {
            background: #f0f0f0;
            width: 12px;
            margin: 0px 0px 0px 0px;
            border-radius: 6px;
        }
        QScrollBar::handle:vertical {
            background: #7a7a7a;
            min-height: 20px;
            border-radius: 6px;
        }
        QScrollBar::handle:vertical:hover {
            background: #555555;
        }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            background: none;
            height: 0px;
        }

        QScrollBar:horizontal {
            background: #f0f0f0;
            height: 12px;
            margin: 0px 0px 0px 0px;
            border-radius: 6px;
        }
        QScrollBar::handle:horizontal {
            background: #7a7a7a;
            min-width: 20px;
            border-radius: 6px;
        }
        QScrollBar::handle:horizontal:hover {
            background: #555555;
        }
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
            background: none;
            width: 0px;
        }
    """)



        self.email = email

        main_lay = self.frameLayout()
        main_lay.setSpacing(21)

        top_lay = QHBoxLayout()
        main_lay.addLayout(top_lay)

        # Add simplified account balance card at the top
        balance_frame = self.create_balance_account(email)
        main_lay.addWidget(balance_frame)

        content_lay = QGridLayout()
        content_lay.setSpacing(24)
        main_lay.addLayout(content_lay)

        small_cards_lay = QVBoxLayout()
        small_cards_lay.setSpacing(17)
        content_lay.addLayout(small_cards_lay, 0, 0)

        stock_symbols = ["AAPL", "TSLA", "MSFT" , "AMZN", "NFLX", "AMD"]

        small_card_details = []

        for symbol in stock_symbols:
            info = get_details(symbol)
            if not info:
                continue  # אם לא חזר מידע, מדלגים

            try:
                change_percent = float(info['change_percent'])
            except (KeyError, ValueError):
                change_percent = 0

            # קביעת צבע וסמל
            if change_percent < 0:
                color = "red"
                icon = "⭝"
            elif change_percent > 0:
                color = "green"
                icon = "⭜"
            else:
                color = "black"
                icon = "⭯"

            # יצירת כרטיס לכל מניה
            card = dict(
                title=f"{symbol}",
                value=(
                    f"<a style='font-size: 24px; font-weight: bold; color: {color};'>"
                    f"{icon} {change_percent:.2f}%</a><br>"
                    f"<a style='font-size: 18px; color: rgba(0, 0, 0, .6);'>"
                    f"Price: ${info['price']}</a>"
                ),
                default_value_style=False
            )

            small_card_details.append(card)

        row, col = 0, 0
        for index, small_card_detail in enumerate(small_card_details):
            if index and not index % 3:
                row += 1
                col = 0

            if not col:
                _small_cards_lay = QHBoxLayout()
                small_cards_lay.addLayout(_small_cards_lay)

            small_card = SmallCard(**small_card_detail)
            _small_cards_lay.addWidget(small_card)

            col += 1

        # Create purchase history table and place it in the right side where ActivityCard was
        self.purchase_table = self.create_purchase_table()

        self.refresh_purchase_table()

        # Create a frame to hold the table with a title
        table_frame = QFrame()
        table_layout = QVBoxLayout(table_frame)

        # Add title to the table frame
        table_title = QLabel("Purchase History")
        table_title.setObjectName("section_header")
        table_layout.addWidget(table_title)

        # Add the table to the frame
        table_layout.addWidget(self.purchase_table)

        # Add the table frame to the content layout where ActivityCard was
        content_lay.addWidget(table_frame, 0, 1)
        content_lay.setColumnStretch(1, 2)
        
        # Add profit/loss graph section with header
        graph_section = QLabel("Annual Performance Analysis")
        graph_section.setObjectName("section_header")
        main_lay.addWidget(graph_section)

        # Create profit/loss graph
        self.profit_loss_graph = self.create_profit_loss_graph()
        main_lay.addWidget(self.profit_loss_graph)

        main_lay.addStretch()



    


    def create_user_profile(self):
        """Create a user profile section displaying the email"""
        profile_frame = QFrame()
        profile_frame.setObjectName("user_profile_frame")

        # Main layout
        layout = QHBoxLayout(profile_frame)

        # User avatar (circle with initials or icon)
        avatar = QFrame()
        avatar.setFixedSize(40, 40)
        avatar.setObjectName("user_avatar")
        avatar_layout = QVBoxLayout(avatar)
        avatar_layout.setContentsMargins(0, 0, 0, 0)

        # Get initials from email
        initials = ''.join([c[0].upper() for c in self.email.split('@')[0].split('.')])[:2]
        initials_label = QLabel(initials)
        initials_label.setObjectName("user_initials")
        initials_label.setAlignment(Qt.AlignCenter)
        avatar_layout.addWidget(initials_label)

        layout.addWidget(avatar)

        # User email and welcome message
        user_info = QVBoxLayout()
        welcome = QLabel(userStatus(self.email))
        welcome.setObjectName("welcome_label")

        email = QLabel(self.email)
        email.setObjectName("user_email_label")

        user_info.addWidget(welcome)
        user_info.addWidget(email)
        layout.addLayout(user_info)

        # Add settings button
        layout.addStretch()
        settings_btn = QPushButton("Settings")
        settings_btn.setObjectName("user_settings_button")
        layout.addWidget(settings_btn)

        return profile_frame
        
    def create_balance_account(self,email):
        """Create a simplified account balance display with just the header and amount"""
        # Create a container frame
        balance_frame = QFrame()
        balance_frame.setObjectName("balance_account_frame")
        balance_frame.setStyleSheet("""
            QFrame#balance_account_frame {
                background-color: white;
                border-radius: 15px;
                margin-bottom: 20px;
                padding: 20px;
            }
        """)

        # Main layout
        layout = QVBoxLayout(balance_frame)

        # Add user email display at the top
        email_layout = QHBoxLayout()

        # User icon - Using a simple text icon instead of SVG for now
        user_icon = QLabel("👤")  # Using a unicode character instead of SVG
        user_icon.setStyleSheet("font-size: 20px;")
        email_layout.addWidget(user_icon)

        # User email label
        email_label = QLabel(userStatus(self.email))
        email_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 500;
            color: #4c4c4c;
            margin-left: 8px;
        """)
        email_layout.addWidget(email_label)

        email_layout.addStretch()
        layout.addLayout(email_layout)

        # Add some space between email and balance
        layout.addSpacing(15)

        # Balance header - using UPPERCASE as requested
        balance_header = QLabel("ACCOUNT BALANCE")
        balance_header.setObjectName("section_header")
        layout.addWidget(balance_header)

        # Current balance amount
        balance = "$" + str(getBalance(self.email))
        balance_amount = QLabel(balance)
        balance_amount.setStyleSheet("""
            font-size: 32px;
            font-weight: bold;
            color: #1B59F8;
            margin-top: 5px;
        """)
        layout.addWidget(balance_amount)

        layout.addStretch()

        return balance_frame

    def create_purchase_table(self):
        # Create table for purchase history
        table = QTableWidget()
        table.setEditTriggers(QTableWidget.NoEditTriggers)

        table.setStyleSheet("""
        QTableWidget {
            background-color: #ffffff;
            border: 1px solid #e0e0e0;
            border-radius: 10px;
            font-size: 16px;
            color: #333;
            gridline-color: #f0f0f0;
            alternate-background-color: #fafafa;
            selection-background-color: #e6f0ff;
            selection-color: #000;
        }

        QHeaderView::section {
            background-color: #f4f6f8;
            color: #2c3e50;
            padding: 8px;
            border: none;
            font-weight: 600;
            font-size: 14px;
        }

        QTableWidget::item {
            padding: 6px;
        }

        QTableWidget::item:selected {
            background-color: #dbeafe;
            color: #000;
        }

        QScrollBar:vertical {
            background: #f0f0f0;
            width: 12px;
            border-radius: 6px;
        }

        QScrollBar::handle:vertical {
            background: #bbb;
            border-radius: 6px;
        }

        QScrollBar::handle:vertical:hover {
            background: #999;
        }

        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            height: 0px;
            background: none;
        }
    """)

        table.setColumnCount(5)

        table.setHorizontalHeaderLabels(["Date", "Stock", "Quantity", "Price", "Total"])

        # קבע מוד קבוע לרוחב העמודות (לא אוטומטי)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)

        # קבע רוחב קבוע לכל עמודה
        fixed_width = 1200
        for col in range(table.columnCount()):
            table.setColumnWidth(col, fixed_width)

        sample_data = format_trades_for_view(self.email)

        if sample_data and sample_data[0][0] == "Error":
            # TODO - replace with message on screen
            print(f"{sample_data[0][1]:}")

        else:
            table.setRowCount(len(sample_data))

            for row, (type_, date, stock, quantity, price, total) in enumerate(sample_data):
                table.setItem(row, 0, QTableWidgetItem(date))
                table.setItem(row, 1, QTableWidgetItem(stock))
                table.setItem(row, 2, QTableWidgetItem(quantity))

                # Highlight current value based on profit/loss
                price_item = QTableWidgetItem(price)
                total_item = QTableWidgetItem(total)

                # Set color based on profit/loss
                if type_ == 1:  # 0 for buy, 1 for sell
                    price_item.setForeground(QColor("green"))
                    total_item.setForeground(QColor("green"))
                else:
                    price_item.setForeground(QColor("red"))
                    total_item.setForeground(QColor("red"))

                table.setItem(row, 3, price_item)
                table.setItem(row, 4, total_item)

            # קביעת רוחב ספציפי לכל עמודה (ברוחב פיקסלים)
            table.setColumnWidth(0, 1200)  # עמודה ראשונה - תאריך
            table.setColumnWidth(1, 100)  # עמודה שניה - מניה
            table.setColumnWidth(2, 80)  # עמודה שלישית - כמות
            table.setColumnWidth(3, 100)  # עמודה רביעית - מחיר
            table.setColumnWidth(4, 120)  # עמודה חמישית - סה"כ

            # הגדרת גובה אחיד לכל השורות
            row_height = 38  # שנה את הערך הזה לגובה הרצוי
            for row in range(table.rowCount()):
                table.setRowHeight(row, row_height)

            # Set table properties
            table.setAlternatingRowColors(True)
            table.setSortingEnabled(True)
            table.resizeColumnsToContents()
            table.setSelectionBehavior(QAbstractItemView.SelectRows)

            # Set a fixed height to match the space available
            table.setMinimumHeight(300)

        return table

    def get_updated_purchase_data(self):
        sample_data = format_trades_for_view(self.email)

        if sample_data and sample_data[0][0] == "Error":
            print(f"{sample_data[0][1]}")
            return []

        return sample_data

    def refresh_purchase_table(self):
        self.purchase_table.clearContents()
        self.purchase_table.setRowCount(0)

        updated_data = self.get_updated_purchase_data()
        self.purchase_table.setRowCount(len(updated_data))

        for row, (type_, date, stock, quantity, price, total) in enumerate(updated_data):
            self.purchase_table.setItem(row, 0, QTableWidgetItem(date))
            self.purchase_table.setItem(row, 1, QTableWidgetItem(stock))
            self.purchase_table.setItem(row, 2, QTableWidgetItem(quantity))

            price_item = QTableWidgetItem(price)
            total_item = QTableWidgetItem(total)

            # Color profit/loss
            if type_ == 1:
                price_item.setForeground(QColor("green"))
                total_item.setForeground(QColor("green"))
            else:
                price_item.setForeground(QColor("red"))
                total_item.setForeground(QColor("red"))

            self.purchase_table.setItem(row, 3, price_item)
            self.purchase_table.setItem(row, 4, total_item)

        # Set row height
        row_height = 38
        for row in range(self.purchase_table.rowCount()):
            self.purchase_table.setRowHeight(row, row_height)

    def create_profit_loss_graph(self):
        # Create a frame to hold the graph
        frame = QFrame()
        frame.setObjectName("graph_frame")
        frame.setStyleSheet("""
            QFrame#graph_frame {
                background-color: white;
                border-radius: 12px;
                padding: 15px;
            }
        """)
        
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Create a title for the annual view
        year_view_title = QLabel("Annual Performance Overview")
        year_view_title.setObjectName("graph_title")
        year_view_title.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #333;
            margin-bottom: 12px;
        """)
        layout.addWidget(year_view_title)
        
        # Create matplotlib figure and canvas with improved dimensions
        self.figure = Figure(figsize=(10, 5), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMinimumHeight(350)  # Increased height with the extra space
        layout.addWidget(self.canvas)
        
        # Plot the annual data by default
        self.plot_data(timeframe="Year")
        
        return frame
    
    def plot_data(self, timeframe="Year"):
        """Plot profit/loss data with enhanced visuals for annual view"""
        # Clear the figure for new plot
        self.figure.clear()
        
        # Create subplot with adjusted margins
        ax = self.figure.add_subplot(111)
        plt.subplots_adjust(left=0.09, right=0.96, top=0.92, bottom=0.15)  # Increased bottom margin for info panel
        
        # Get data from the purchase table
        dates = []
        balance_changes = []
        running_balance = []
        transaction_details = []  # Store details for tooltips
        current_balance = 0
        
        # Extract data from the table
        rows = self.purchase_table.rowCount()
        for row in range(rows):
            try:
                date_str = self.purchase_table.item(row, 0).text()
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                
                # Set time values proportionally to avoid overlapping points
                date = date.replace(hour=row % 24, minute=(row * 13) % 60)
                
                # Get the total amount and transaction description
                total_text = self.purchase_table.item(row, 4).text()
                stock = self.purchase_table.item(row, 1).text() if self.purchase_table.item(row, 1) else "Unknown"
                quantity = self.purchase_table.item(row, 2).text() if self.purchase_table.item(row, 2) else "0"
                description = f"{stock} ({quantity})"
                
                # Remove $ and commas to convert to float
                total_value = float(total_text.replace('$', '').replace(',', '').strip())
                
                # Determine if it's a buy (negative) or sell (positive)
                if self.purchase_table.item(row, 4).foreground().color().name() == "#ff0000":
                    # Buy operation (negative impact on balance)
                    total_value = -total_value
                    transaction_type = "Purchase"
                else:
                    transaction_type = "Sale"
                
                dates.append(date)
                balance_changes.append(total_value)
                
                # Update running balance
                current_balance += total_value
                running_balance.append(current_balance)
                
                # Store details for tooltips
                transaction_details.append({
                    'description': description,
                    'amount': total_value,
                    'balance': current_balance,
                    'type': transaction_type,
                    'stock': stock,
                    'quantity': quantity
                })
                
            except (ValueError, AttributeError) as e:
                print(f"Error processing row {row}: {e}")
        
        # Filter data based on timeframe - always annual view now
        if not dates:
            # No data to display
            ax.text(0.5, 0.5, "No transaction data available", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=12)
            self.canvas.draw()
            return
        
        # Set cutoff to one year
        today = datetime.datetime.now()
        cutoff = today - timedelta(days=365)
        
        # Filter transactions
        filtered_dates = []
        filtered_balance = []
        filtered_details = []
        for i, date in enumerate(dates):
            if date >= cutoff:
                filtered_dates.append(date)
                filtered_balance.append(running_balance[i])
                filtered_details.append(transaction_details[i])
                
        # Plot data
        if filtered_dates:
            # Use a more modern style
            plt.style.use('seaborn-v0_8-whitegrid')
            
            # Stock market colors
            line_color = '#0F4C81'  # Dark blue for lines (traditional stock chart color)
            positive_color = '#2E8B57'  # Dark green for gains
            negative_color = '#B22222'  # Dark red for losses
            marker_color = '#333333'  # Darker marker color
            
            # Create scatter plot for interactive data points with improved visuals
            scatter = ax.scatter(filtered_dates, filtered_balance, 
                    s=45,  # Marker size
                    color=marker_color,
                    alpha=0.8,
                    edgecolor='white',
                    linewidth=1,
                    zorder=5)
            
            # Create smoother line with gradient
            line = ax.plot(filtered_dates, filtered_balance, 
                    color=line_color, linewidth=2.0, 
                    alpha=0.85, zorder=4)[0]
            
            # Add horizontal line at y=0 with better styling
            ax.axhline(y=0, color='#888888', linestyle='-', alpha=0.4, linewidth=1)
            
            # Fill area above/below zero line with stock market colors
            ax.fill_between(filtered_dates, filtered_balance, 0, 
                        where=[b > 0 for b in filtered_balance], 
                        color=positive_color, alpha=0.25, interpolate=True)
            ax.fill_between(filtered_dates, filtered_balance, 0, 
                        where=[b < 0 for b in filtered_balance], 
                        color=negative_color, alpha=0.25, interpolate=True)
            
            # Make the axes less prominent
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_alpha(0.3)
            ax.spines['bottom'].set_alpha(0.3)
            
            # Set compact tick format for y-axis
            import matplotlib.ticker as ticker
            def currency_formatter(x, pos):
                if abs(x) >= 1000000:
                    return f'${x/1000000:.1f}M'
                elif abs(x) >= 1000:
                    return f'${x/1000:.0f}K'
                return f'${x:.0f}'
            
            ax.yaxis.set_major_formatter(ticker.FuncFormatter(currency_formatter))
            
            # Use month locator for year view
            import matplotlib.dates as mdates
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
            
            # Rotate date labels for better readability
            plt.setp(ax.get_xticklabels(), rotation=30, ha='right', rotation_mode='anchor', fontsize=10)
            plt.setp(ax.get_yticklabels(), fontsize=10)
            
            # Add monthly average line
            # Group by month and calculate average
            months = {}
            for i, date in enumerate(filtered_dates):
                month_key = date.strftime('%Y-%m')
                if month_key not in months:
                    months[month_key] = []
                months[month_key].append(filtered_balance[i])
            
            monthly_avg = []
            monthly_dates = []
            for month, values in sorted(months.items()):
                avg = sum(values) / len(values)
                monthly_avg.append(avg)
                # Use middle of month for x position
                year, month = month.split('-')
                monthly_dates.append(datetime.datetime(int(year), int(month), 15))
            
            if len(monthly_dates) > 1:
                # Add monthly average line
                ax.plot(monthly_dates, monthly_avg, 
                        color='#5D5D5D', linewidth=1.5, 
                        linestyle='--', alpha=0.7,
                        label='Monthly Average')
                
                # Add legend with more professional style
                legend = ax.legend(loc='upper left', frameon=True, framealpha=0.85, fontsize=9)
                legend.get_frame().set_edgecolor('#CCCCCC')
            
            # Add min/max markers - more subtle
            if filtered_balance:
                max_balance = max(filtered_balance)
                min_balance = min(filtered_balance)
                max_idx = filtered_balance.index(max_balance)
                min_idx = filtered_balance.index(min_balance)
                
                # Add markers for min/max with stock market colors
                ax.scatter([filtered_dates[max_idx]], [max_balance], 
                           s=90, color=positive_color, edgecolor='white', 
                           linewidth=1.2, zorder=6, marker='o')
                ax.scatter([filtered_dates[min_idx]], [min_balance], 
                           s=90, color=negative_color, edgecolor='white', 
                           linewidth=1.2, zorder=6, marker='o')
                
                # Add annotations for min/max - more subtle
                ax.annotate(f"${max_balance:,.2f}", 
                            xy=(filtered_dates[max_idx], max_balance),
                            xytext=(0, 15), textcoords='offset points',
                            ha='center', fontsize=9, fontweight='bold',
                            bbox=dict(boxstyle="round,pad=0.3", fc=positive_color, ec="white", alpha=0.7))
                
                ax.annotate(f"${min_balance:,.2f}", 
                            xy=(filtered_dates[min_idx], min_balance),
                            xytext=(0, -25), textcoords='offset points',
                            ha='center', fontsize=9, fontweight='bold',
                            bbox=dict(boxstyle="round,pad=0.3", fc=negative_color, ec="white", alpha=0.7))
            
            # Calculate YTD change
            if filtered_balance:
                start_balance = filtered_balance[0] if filtered_balance else 0
                end_balance = filtered_balance[-1] if filtered_balance else 0
                change = end_balance - start_balance
                pct_change = (change / abs(start_balance)) * 100 if start_balance != 0 else 0
                
                # Create YTD change text
                stats_text = f"YTD Change: "
                stats_text += f"${change:+,.2f}" if change != 0 else "$0.00"
                stats_text += f" ({pct_change:+.2f}%)" if pct_change != 0 else " (0.00%)"
                
                # Add YTD change in the top right corner instead of on the graph
                # Use color based on positive/negative
                text_color = positive_color if change >= 0 else negative_color
                
                # Position the YTD change in top right with appropriate coloring
                ax.text(0.98, 0.98, stats_text,
                        transform=ax.transAxes,
                        fontsize=10, fontweight='bold',
                        color=text_color,
                        ha='right', va='top',
                        bbox=dict(boxstyle="round,pad=0.4", fc="white", ec="#DDDDDD", alpha=0.9))
            
            
            # Create a fixed info panel in bottom left instead of floating tooltips
            info_panel = ax.text(0.01, 0.01, "Hover over points to see details",
                                transform=ax.transAxes,
                                fontsize=10, 
                                ha='left', va='bottom',
                                bbox=dict(boxstyle="round,pad=0.4", fc="#F8F9FA", ec="#CCCCCC", alpha=0.95))
            
            # Setup hover functionality with fixed position tooltip
            def hover(event):
                if event.inaxes == ax:
                    cont, ind = scatter.contains(event)
                    if cont:
                        # Get data point index
                        idx = ind["ind"][0]
                        if idx < len(filtered_dates):
                            # Format tooltip text with enhanced layout
                            detail = filtered_details[idx]
                            date_str = filtered_dates[idx].strftime("%Y-%m-%d")
                            text = f"{detail['stock']} - {detail['type']}\n"
                            text += f"Date: {date_str}\n"
                            text += f"Quantity: {detail['quantity']}\n"
                            
                            # Use appropriate colors for purchases/sales
                            if detail['type'] == "Purchase":
                                amount_text = f"Amount: -${abs(detail['amount']):,.2f}"
                            else:
                                amount_text = f"Amount: +${abs(detail['amount']):,.2f}"
                                
                            text += f"{amount_text}\n"
                            text += f"Balance: ${detail['balance']:,.2f}"
                            
                            # Update the fixed info panel
                            info_panel.set_text(text)
                            self.canvas.draw_idle()
                    else:
                        # Reset info panel when not hovering over a point
                        info_panel.set_text("Hover over points to see details")
                        self.canvas.draw_idle()
                            
            # Connect hover event
            self.canvas.mpl_connect("motion_notify_event", hover)
            
            # Add subtle watermark with alpha
            ax.text(0.98, 0.01, 'Finance Tracker Pro', 
                    transform=ax.transAxes, 
                    fontsize=8, color='gray', alpha=0.4,
                    ha='right', va='bottom')
            
        else:
            # No data for the selected timeframe
            ax.text(0.5, 0.5, "No data available for annual timeframe", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=11)
        
        # Adjust layout to make sure everything fits
        self.figure.tight_layout()
        
        # Update the canvas
        self.canvas.draw()

class Body(QFrame):
    def __init__(self,email):
        super().__init__()
        
        main_lay = QVBoxLayout(self)

        top_lay = QHBoxLayout()
        main_lay.addLayout(top_lay)

        reports = QLabel("Reports")
        reports.setObjectName("reports")
        top_lay.addWidget(reports)

        top_lay.addStretch()

        download_button = QPushButton("Download")
        download_button.setIcon(QIcon("./View/svgs/download.svg"))  # תוודא שהנתיב נכון
        download_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                font-size: 14px;
                color: #4c4c4c;
            }
            QPushButton:hover {
                color: #1B59F8; 
            }
        """)
        
        top_lay.addWidget(download_button)

        # חיבור לפונקציה שתיצור את קובץ האקסל
        download_button.clicked.connect(lambda: self.export_table_to_excel(body_content_frame.purchase_table))


        main_lay.addSpacing(37)

        top_hline = HLine()
        main_lay.addWidget(top_hline)

        main_lay.addSpacing(28)

        self.body_content_frame = BodyContentFrame(email)
        main_lay.addWidget(self.body_content_frame, 1)

        main_lay.addStretch()

    def refresh_table(self):
        print("Refreshing Body table")

        self.body_content_frame.refresh_purchase_table()

    def export_table_to_excel(self, table):
        downloads_folder = str(Path.home() / "Downloads")
        file_path = os.path.join(downloads_folder, "exported_table.xlsx")

        wb = Workbook()
        ws = wb.active

        # כתיבת כותרות
        for col in range(table.columnCount()):
            header = table.horizontalHeaderItem(col)
            ws.cell(row=1, column=col+1, value=header.text() if header else "")

        # כתיבת שורות
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                ws.cell(row=row+2, column=col+1, value=item.text() if item else "")

        wb.save(file_path)
        print(f"download succes {file_path}")

if __name__ == "__main__":
    import os

    os.system("python main.py")